import os
import openpyxl
import pandas as pd

def update_excel_with_csv(csv_path, excel_path, sheet_name):
    csv_df = pd.read_csv(csv_path)
    # exist then append mode, or write mode
    mode = "a" if os.path.exists(excel_path) else "w"

    with pd.ExcelWriter(excel_path, mode=mode, engine="openpyxl") as writer:
        csv_df.to_excel(writer, sheet_name=sheet_name, index=False)


def concatenate_csv_files(files:list, output_file:str):
    # Read the two CSV files into DataFrames
    dfs = []
    for file in files:
        dfs.append(pd.read_csv(file))
    
    # Concatenate the DataFrames
    concatenated_df = pd.concat(dfs, ignore_index=True)
    
    # Save the concatenated DataFrame to a new CSV file
    concatenated_df.to_csv(output_file, index=False)
    print(f'Concatenated file saved as {output_file}')


def adjust_column(series):
    if series.empty:
        return series

    offset = 0
    adjusted = []
    
    last_raw = series.iloc[0]
    last_adjusted = series.iloc[0]
    adjusted.append(last_adjusted)
    
    for raw in series.iloc[1:]:
        if raw < last_raw:
            offset = last_adjusted
        new_val = raw + offset
        adjusted.append(new_val)
        
        last_raw = raw
        last_adjusted = new_val
    
    return adjusted
 
 
def prep_pellet_count(path: str):
    """when combining two csv files, making the pellet count column increasing
    instead of go to 0 for the new file.

    Args:
        path (str): path of combined csv files
    """
    df = pd.read_csv(path)
    columns_to_adjust = ["Pellet_Count", "Left_Poke_Count", "Right_Poke_Count"]
    for col in columns_to_adjust:
        df[col] = adjust_column(df[col])
    df.to_csv(path[:-4]+'_1.CSV', index=False)


def check_data_by_date(path:str):
    df = pd.read_csv(path)
    df['MM:DD:YYYY hh:mm:ss'] = pd.to_datetime(df['MM:DD:YYYY hh:mm:ss'])

    index_drop = (df['MM:DD:YYYY hh:mm:ss'].diff() < pd.Timedelta(0)).idxmax()
    df_part1 = df.iloc[:index_drop]
    df_part2 = df.iloc[index_drop:]

    df_corrected = pd.concat([df_part2, df_part1]).reset_index(drop=True)
    df_corrected.to_csv(path, index=False)


def merge_csvs_to_excel(excel:str, csv_root:str):
    files = os.listdir(csv_root)
    if '.DS_Store' in files: 
        files.remove('.DS_Store')

    # construct group meta
    for file in files:
        path = os.path.join(csv_root, file)
        print(path)
        name = file.split('.')[0]
        sheet = name[:len(name)-2]+'.'+name[-2:]
        update_excel_with_csv(csv_path=path, excel_path=excel, sheet_name=sheet)


def split_ctrl_cask(excel_path:str):
    xls = pd.ExcelFile(excel_path)  
    all_sheets = xls.sheet_names
    ctrl_sheets = dict()
    cask_sheets = dict()
    for sheet in all_sheets:
        if sheet.startswith('B'): 
            group = int(sheet.split('.')[0][1:])
            if group in [2, 3, 5, 7, 11, 12]:
                ctrl_sheets[sheet] = xls.parse(sheet)
            else:
                cask_sheets[sheet] = xls.parse(sheet)
        elif sheet.startswith('C'):
            if sheet in ['C1.M1', 'C1.M2', 'C2.M1', 'C2.M3', 'C3.M4']:
                ctrl_sheets[sheet] = xls.parse(sheet)
            else:
                cask_sheets[sheet] = xls.parse(sheet)

    with pd.ExcelWriter("../FR1_ctrl.xlsx", engine="openpyxl") as writer:
        for sheet_name, data in ctrl_sheets.items():
            data.to_excel(writer, sheet_name=sheet_name, index=False)

    with pd.ExcelWriter("../FR1_cask.xlsx", engine="openpyxl") as writer:
        for sheet_name, data in cask_sheets.items():
            data.to_excel(writer, sheet_name=sheet_name, index=False)


def copy_sheet(input_file, target_file, sheet_name):
    try:
        input_wb = openpyxl.load_workbook(input_file, data_only=True)
        if sheet_name not in input_wb.sheetnames:
            print(f"Sheet '{sheet_name}' does not exist in {input_file}")
            return

        input_sheet = input_wb[sheet_name]
        target_wb = openpyxl.load_workbook(target_file)

        if sheet_name in target_wb.sheetnames:
            print(f"Sheet '{sheet_name}' already exists in {target_file}.")
            return

        target_sheet = target_wb.create_sheet(title=sheet_name.strip())
        for row in input_sheet.iter_rows():
            for cell in row:
                target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)

        target_wb.save(target_file)
        print(f"Copied sheet '{sheet_name}' from {input_file} to {target_file}.")
    except Exception as e:
        print(f"Error copying sheet: {e}")


if __name__ == '__main__':
    # merge_csvs_to_excel(excel='../reversal_cask.xlsx', csv_root='../CASK_Data/reversal/cask')
    # split_ctrl_cask('../FR1_collection.xlsx')
    # check_data_by_date('../reversal_ctrl.xlsx', 'C1.M1')
    # copy_sheet('../reversal_cask.xlsx', '../reversal_ctrl.xlsx', 'C5.M2')
    # file_1 = '../new_data/M1/FED000_022625_01.CSV'
    # file_2 = '../new_data/M1/FED000_022625_03.CSV'
    # concatenate_csv_files([file_1, file_2], '../new_data/M1/FED000_022625_04.CSV')
    # prep_pellet_count('../Food Intake Data/M5_fen.csv')
    update_excel_with_csv(excel_path='../data/reversal_female.xlsx', csv_path='../new_data/M1/FED000_022625_04.CSV', sheet_name='R3M1')
    update_excel_with_csv(excel_path='../data/reversal_female.xlsx', csv_path='../new_data/M2/FED000_022625_01.CSV', sheet_name='R3M2')
    update_excel_with_csv(excel_path='../data/reversal_female.xlsx', csv_path='../new_data/M3/FED000_022625_01.CSV', sheet_name='R3M3')
    update_excel_with_csv(excel_path='../data/reversal_female.xlsx', csv_path='../new_data/M4/FED000_022625_01.CSV', sheet_name='R3M4')